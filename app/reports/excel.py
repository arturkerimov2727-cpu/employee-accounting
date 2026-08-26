import io

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


BLUE = "1D67CF"


def column_letter(number):
    result = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result


def total_formula(column, first_row, last_row):
    if last_row < first_row:
        return "=0"
    return f"=SUM('Посещаемость'!{column}{first_row}:{column}{last_row})"


def style_header(cells):
    for cell in cells:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def create_excel_report(period, headings, values):
    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    sheet = workbook.active
    sheet.title = "Посещаемость"
    last_column = len(headings) + 1
    first_data_row = 3
    last_data_row = first_data_row + len(values) - 1

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    sheet.cell(1, 1, f"Отчёт посещаемости за период: {period}")
    sheet.cell(1, 1).font = Font(bold=True, size=14, color="FFFFFF")
    sheet.cell(1, 1).fill = PatternFill("solid", fgColor=BLUE)
    sheet.cell(1, 1).alignment = Alignment(horizontal="center")

    for column, heading in enumerate([*headings, "Рабочие часы"], start=1):
        sheet.cell(2, column, heading)
    style_header(sheet[2])

    for row_number, value in enumerate(values, start=first_data_row):
        for column, item in enumerate(value, start=1):
            sheet.cell(row_number, column, item)
        sheet.cell(row_number, last_column, f"=D{row_number}/60")
        sheet.cell(row_number, last_column).number_format = "0.0"

    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{column_letter(last_column)}{max(2, last_data_row)}"
    if values:
        table = Table(displayName="AttendanceReport", ref=f"A2:{column_letter(last_column)}{last_data_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)

    for column in range(1, last_column + 1):
        letter = column_letter(column)
        width = max(len(str(sheet.cell(row, column).value or "")) for row in range(1, max(2, last_data_row) + 1)) + 2
        sheet.column_dimensions[letter].width = min(30, max(12, width))
    sheet.row_dimensions[1].height = 26

    summary = workbook.create_sheet("Сводка")
    summary.merge_cells("A1:B1")
    summary["A1"] = f"Сводка за период: {period}"
    summary["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor=BLUE)
    summary["A1"].alignment = Alignment(horizontal="center")
    summary["A3"] = "Показатель"
    summary["B3"] = "Значение"
    style_header(summary[3][:2])

    metrics = [
        ("Сотрудников в отчёте", f"=COUNTA('Посещаемость'!A{first_data_row}:A{last_data_row})" if values else "=0"),
        ("Всего смен", total_formula("C", first_data_row, last_data_row)),
        ("Рабочих часов", f"=SUM('Посещаемость'!{column_letter(last_column)}{first_data_row}:{column_letter(last_column)}{last_data_row})" if values else "=0"),
        ("Опозданий", total_formula("E", first_data_row, last_data_row)),
        ("Минут опозданий", total_formula("F", first_data_row, last_data_row)),
        ("Ранних уходов", total_formula("G", first_data_row, last_data_row)),
        ("Часов переработки", f"={total_formula('H', first_data_row, last_data_row)[1:]}/60"),
        ("Отсутствий", total_formula("I", first_data_row, last_data_row)),
    ]
    for row_number, (label, formula) in enumerate(metrics, start=4):
        summary.cell(row_number, 1, label)
        summary.cell(row_number, 2, formula)
        summary.cell(row_number, 2).number_format = "0.0" if row_number in {6, 10} else "0"

    summary["A14"] = "Отдел"
    summary["B14"] = "Рабочие часы"
    style_header(summary[14][:2])
    departments = sorted({str(row[1] or "Без отдела") for row in values})
    for row_number, department in enumerate(departments, start=15):
        summary.cell(row_number, 1, department)
        summary.cell(
            row_number,
            2,
            f'=SUMIF(\'Посещаемость\'!$B${first_data_row}:$B${last_data_row},A{row_number},\'Посещаемость\'!${column_letter(last_column)}${first_data_row}:${column_letter(last_column)}${last_data_row})',
        )
        summary.cell(row_number, 2).number_format = "0.0"

    summary["D3"] = "Сотрудник"
    summary["E3"] = "Рабочие часы"
    style_header(summary[3][3:5])
    for row_number, source_row in enumerate(range(first_data_row, last_data_row + 1), start=4):
        summary.cell(row_number, 4, f"='Посещаемость'!A{source_row}")
        summary.cell(row_number, 5, f"='Посещаемость'!{column_letter(last_column)}{source_row}")
        summary.cell(row_number, 5).number_format = "0.0"

    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 18
    summary.column_dimensions["D"].width = 32
    summary.column_dimensions["E"].width = 18
    summary.freeze_panes = "A4"

    if values:
        employee_chart = BarChart()
        employee_chart.type = "bar"
        employee_chart.style = 10
        employee_chart.title = "Рабочие часы по сотрудникам"
        employee_chart.y_axis.title = "Сотрудники"
        employee_chart.x_axis.title = "Часы"
        employee_chart.legend = None
        employee_chart.height = 8
        employee_chart.width = 16
        employee_chart.add_data(Reference(summary, min_col=5, min_row=3, max_row=last_data_row + 1), titles_from_data=True)
        employee_chart.set_categories(Reference(summary, min_col=4, min_row=4, max_row=last_data_row + 1))
        summary.add_chart(employee_chart, "G3")

        department_chart = PieChart()
        department_chart.style = 10
        department_chart.title = "Доля рабочих часов по отделам"
        department_chart.height = 8
        department_chart.width = 14
        department_chart.add_data(Reference(summary, min_col=2, min_row=14, max_row=14 + len(departments)), titles_from_data=True)
        department_chart.set_categories(Reference(summary, min_col=1, min_row=15, max_row=14 + len(departments)))
        summary.add_chart(department_chart, "G20")

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
